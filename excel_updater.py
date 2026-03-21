import sys
import subprocess
import os
import json

# Ensure openpyxl is installed
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, '-m', 'pip', 'install', 'openpyxl'])
    import openpyxl
from openpyxl import Workbook, load_workbook

json_file = os.path.join(os.path.dirname(__file__), 'files_to_update.json')
excel_file = os.path.join(os.path.dirname(__file__), '0.FilesUpdate.xlsx')

with open(json_file, 'r', encoding='utf-8') as f:
    data = json.load(f)

if os.path.exists(excel_file):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
    except Exception:
        wb = Workbook()
        ws = wb.active
        ws.append(["일시", "파일명", "요청 요약"])
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["일시", "파일명", "요청 요약"])

for item in data:
    ws.append([item.get("일시", ""), item.get("파일명", ""), item.get("요청 요약", "")])

wb.save(excel_file)
print("Excel file successfully created and updated with openpyxl.")
